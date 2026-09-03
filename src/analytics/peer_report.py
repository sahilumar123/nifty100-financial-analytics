from pathlib import Path
import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "data" / "database" / "nifty100.db"
OUTPUT = ROOT / "output" / "peer_comparison.xlsx"


def create_report():

    conn = sqlite3.connect(DB_PATH)

    data = pd.read_sql(
        """
        SELECT
            p.*,
            c.company_name
        FROM peer_percentiles p
        LEFT JOIN companies c
            ON p.company_id = c.id
        """,
        conn,
    )

    conn.close()

    OUTPUT.parent.mkdir(exist_ok=True)

    with pd.ExcelWriter(
        OUTPUT,
        engine="openpyxl",
    ) as writer:

        for group, group_df in data.groupby(
            "peer_group_name"
        ):

            if group == "No peer group assigned":
                continue

            pivot = group_df.pivot_table(
                index=[
                    "company_id",
                    "company_name",
                    "year",
                ],
                columns="metric",
                values=[
                    "value",
                    "percentile_rank",
                ],
            )

            pivot.columns = [
                "_".join(map(str, col))
                for col in pivot.columns
            ]

            pivot.reset_index().to_excel(
                writer,
                sheet_name=group[:31],
                index=False,
            )

    wb = load_workbook(OUTPUT)

    green = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    yellow = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )

    red = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    for ws in wb.worksheets:

        for row in ws.iter_rows(
            min_row=2
        ):

            for cell in row:

                if (
                    isinstance(cell.value, float)
                    and 0 <= cell.value <= 1
                    and "percentile" in str(
                        ws.cell(
                            row=1,
                            column=cell.column,
                        ).value
                    ).lower()
                ):

                    if cell.value >= 0.75:
                        cell.fill = green

                    elif cell.value >= 0.25:
                        cell.fill = yellow

                    else:
                        cell.fill = red

    wb.save(OUTPUT)

    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    create_report()