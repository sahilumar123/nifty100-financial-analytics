from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from src.screener.engine import run_all_presets


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = ROOT / "output" / "screener_output.xlsx"


def export_screeners():

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    results = run_all_presets()

    with __import__("pandas").ExcelWriter(
        OUTPUT,
        engine="openpyxl",
    ) as writer:

        for preset, df in results.items():

            export_df = df.copy()

            export_df.to_excel(
                writer,
                sheet_name=preset[:31],
                index=False,
            )

    wb = load_workbook(OUTPUT)

    green = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    red = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    for ws in wb.worksheets:

        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
        ):

            for cell in row:

                if isinstance(cell.value, (int, float)):

                    if cell.value > 0:
                        cell.fill = green
                    else:
                        cell.fill = red

    wb.save(OUTPUT)

    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    export_screeners()